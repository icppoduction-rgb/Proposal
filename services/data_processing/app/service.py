from __future__ import annotations

import csv
import json
import uuid
from datetime import UTC, date, datetime
from itertools import zip_longest
from math import ceil
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from pyarrow import Table, array, schema as arrow_schema
from pyarrow import types as pa_types
from pyarrow import parquet as pq

from cybersec_platform.contracts.api import (
    DatasetFormat,
    DatasetManifest,
    EditorPageOut,
    EditorRowOut,
    EditorSaveOut,
    EditorSessionOut,
    FeatureFamily,
    SourceType,
)
from cybersec_platform.db.session import get_settings
from cybersec_platform.ml.normalization import (
    ContractValidationError,
    NormalizationEngine,
    UnsupportedDatasetFormatError,
    detect_dataset_format,
    iter_system_call_records,
)

DEFAULT_SHEET_KEY = "__default__"


class DataProcessingError(ValueError):
    def __init__(self, message: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.status_code = status_code


def get_editor_root() -> Path:
    root = Path(get_settings().tmp_path) / "data-editor"
    root.mkdir(parents=True, exist_ok=True)
    return root


def get_session_dir(session_id: str) -> Path:
    return get_editor_root() / session_id


def get_session_path(session_id: str) -> Path:
    return get_session_dir(session_id) / "session.json"


def get_raw_root() -> Path:
    root = Path(get_settings().raw_data_path).resolve()
    root.mkdir(parents=True, exist_ok=True)
    return root


def _session_sheet_key(active_sheet: str | None) -> str:
    return active_sheet or DEFAULT_SHEET_KEY


def _load_session(session_id: str) -> dict[str, Any]:
    session_path = get_session_path(session_id)
    if not session_path.exists():
        raise DataProcessingError("Editor session not found", status_code=404)
    return json.loads(session_path.read_text(encoding="utf-8"))


def _save_session(payload: dict[str, Any]) -> dict[str, Any]:
    session_dir = get_session_dir(payload["session_id"])
    session_dir.mkdir(parents=True, exist_ok=True)
    get_session_path(payload["session_id"]).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def _delete_session_dir(session_id: str) -> None:
    session_dir = get_session_dir(session_id)
    if not session_dir.exists():
        return
    for path in sorted(session_dir.rglob("*"), reverse=True):
        if path.is_file():
            path.unlink(missing_ok=True)
        elif path.is_dir():
            try:
                path.rmdir()
            except OSError:
                continue
    try:
        session_dir.rmdir()
    except OSError:
        pass


def _ensure_raw_file_path(file_path: str) -> Path:
    raw_root = get_raw_root()
    target = Path(file_path).resolve()
    if not target.is_relative_to(raw_root):
        raise DataProcessingError("File path must be inside the raw data directory", status_code=403)
    if not target.exists() or not target.is_file():
        raise DataProcessingError("Raw file not found", status_code=404)
    return target


def _deduplicate_headers(values: list[str]) -> list[str]:
    counters: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = (value or f"column_{index}").strip() or f"column_{index}"
        count = counters.get(base, 0)
        counters[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def _jsonable_value(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=UTC).isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _infer_type_label(value: Any) -> str:
    if value is None:
        return "string"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    return "string"


def _is_scalar_value(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _parse_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes"}:
        return True
    if normalized in {"false", "0", "no"}:
        return False
    raise DataProcessingError(f"Unable to parse boolean value: {value}")


def _coerce_value(value: Any, type_label: str, strict: bool = False) -> Any:
    if value in ("", None):
        return None
    try:
        if type_label == "int":
            return int(value)
        if type_label == "float":
            return float(value)
        if type_label == "bool":
            return _parse_bool(value)
        if type_label == "datetime":
            if isinstance(value, datetime):
                return value
            return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if type_label == "date":
            if isinstance(value, date) and not isinstance(value, datetime):
                return value
            return date.fromisoformat(str(value))
        return str(value)
    except (TypeError, ValueError) as exc:
        if strict:
            raise DataProcessingError(f"Unable to coerce value '{value}' to {type_label}") from exc
        return value


def _detect_json_mode(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig") as file:
        while True:
            char = file.read(1)
            if not char:
                return "ndjson"
            if not char.isspace():
                return "array" if char == "[" else "ndjson"


def _read_json_array_records(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, list):
        raise DataProcessingError("JSON editor supports only arrays of object records")
    records: list[dict[str, Any]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise DataProcessingError("JSON editor supports only arrays of object records")
        records.append(_ensure_flat_record(item))
    return records


def _iter_json_ndjson_records(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig") as file:
        for line_number, line in enumerate(file, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            payload = json.loads(stripped)
            if not isinstance(payload, dict):
                raise DataProcessingError(f"JSON line {line_number} is not an object record")
            yield _ensure_flat_record(payload)


def _ensure_flat_record(record: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, (dict, list)):
            raise DataProcessingError("Nested JSON values are not supported for editing")
        normalized[str(key)] = _jsonable_value(value)
    return normalized


def _iter_delimited_rows(path: Path, delimiter: str) -> Iterator[tuple[int, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        header_row = next(reader, None)
        if header_row is None:
            return
        headers = _deduplicate_headers([str(item) for item in header_row])
        for row_index, row in enumerate(reader):
            values = {
                column: _jsonable_value(value)
                for column, value in zip_longest(headers, row, fillvalue=None)
            }
            yield row_index, values


def _build_delimited_state(path: Path, delimiter: str) -> dict[str, Any]:
    total_rows = 0
    headers: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        header_row = next(reader, None)
        if header_row is not None:
            headers = _deduplicate_headers([str(item) for item in header_row])
        for _ in reader:
            total_rows += 1
    return {
        "base_columns": headers,
        "column_types": {column: "string" for column in headers},
        "total_rows": total_rows,
        "deleted_rows": [],
        "deleted_columns": [],
        "cell_patches": {},
    }


def _iter_json_rows(path: Path, mode: str) -> Iterator[tuple[int, dict[str, Any]]]:
    if mode == "array":
        for index, record in enumerate(_read_json_array_records(path)):
            yield index, record
        return
    for index, record in enumerate(_iter_json_ndjson_records(path)):
        yield index, record


def _build_json_state(path: Path, mode: str) -> dict[str, Any]:
    columns: list[str] = []
    column_types: dict[str, str] = {}
    total_rows = 0
    for _, record in _iter_json_rows(path, mode):
        total_rows += 1
        for key, value in record.items():
            if key not in columns:
                columns.append(key)
            if key not in column_types and value is not None:
                column_types[key] = _infer_type_label(value)
    for column in columns:
        column_types.setdefault(column, "string")
    return {
        "base_columns": columns,
        "column_types": column_types,
        "total_rows": total_rows,
        "deleted_rows": [],
        "deleted_columns": [],
        "cell_patches": {},
    }


def _iter_system_call_rows(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    for row_index, record in enumerate(iter_system_call_records(path)):
        yield row_index, {key: _jsonable_value(value) for key, value in record.items()}


def _build_system_call_state(path: Path) -> dict[str, Any]:
    columns: list[str] = []
    column_types: dict[str, str] = {}
    total_rows = 0
    for _, record in _iter_system_call_rows(path):
        total_rows += 1
        for key, value in record.items():
            if key not in columns:
                columns.append(key)
            if key not in column_types and value is not None:
                column_types[key] = _infer_type_label(value)
    for column in columns:
        column_types.setdefault(column, "string")
    return {
        "base_columns": columns,
        "column_types": column_types,
        "total_rows": total_rows,
        "deleted_rows": [],
        "deleted_columns": [],
        "cell_patches": {},
    }


def _iter_parquet_rows(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    parquet_file = pq.ParquetFile(path)
    row_index = 0
    for batch in parquet_file.iter_batches(batch_size=512):
        for row in batch.to_pylist():
            yield row_index, {key: _jsonable_value(value) for key, value in row.items()}
            row_index += 1


def _is_supported_arrow_type(type_name: Any) -> bool:
    return any(
        (
            pa_types.is_integer(type_name),
            pa_types.is_floating(type_name),
            pa_types.is_boolean(type_name),
            pa_types.is_string(type_name),
            pa_types.is_large_string(type_name),
            pa_types.is_timestamp(type_name),
            pa_types.is_date32(type_name),
            pa_types.is_date64(type_name),
            pa_types.is_null(type_name),
        )
    )


def _column_type_label_from_arrow(type_name: Any) -> str:
    if pa_types.is_integer(type_name):
        return "int"
    if pa_types.is_floating(type_name):
        return "float"
    if pa_types.is_boolean(type_name):
        return "bool"
    if pa_types.is_timestamp(type_name):
        return "datetime"
    if pa_types.is_date32(type_name) or pa_types.is_date64(type_name):
        return "date"
    return "string"


def _build_parquet_state(path: Path) -> dict[str, Any]:
    parquet_file = pq.ParquetFile(path)
    schema = parquet_file.schema_arrow
    for field in schema:
        if not _is_supported_arrow_type(field.type):
            raise DataProcessingError(f"Parquet column '{field.name}' uses an unsupported nested type")
    columns = list(schema.names)
    return {
        "base_columns": columns,
        "column_types": {field.name: _column_type_label_from_arrow(field.type) for field in schema},
        "total_rows": parquet_file.metadata.num_rows if parquet_file.metadata else 0,
        "deleted_rows": [],
        "deleted_columns": [],
        "cell_patches": {},
    }


def _iter_xlsx_rows(path: Path, sheet_name: str) -> Iterator[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return
        headers = _deduplicate_headers([str(item) if item is not None else "" for item in header_row])
        for row_index, row in enumerate(rows):
            values = {
                column: _jsonable_value(value)
                for column, value in zip_longest(headers, row, fillvalue=None)
            }
            yield row_index, values
    finally:
        workbook.close()


def _build_xlsx_state(path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        headers = _deduplicate_headers([str(item) if item is not None else "" for item in (header_row or [])])
        column_types = {column: "string" for column in headers}
        total_rows = 0
        for row in rows:
            total_rows += 1
            for column, value in zip_longest(headers, row, fillvalue=None):
                if column_types[column] == "string" and value is not None:
                    column_types[column] = _infer_type_label(value)
        return {
            "base_columns": headers,
            "column_types": column_types,
            "total_rows": total_rows,
            "deleted_rows": [],
            "deleted_columns": [],
            "cell_patches": {},
        }
    finally:
        workbook.close()


def _build_pcap_projection(path: Path, output_path: Path) -> None:
    manifest = DatasetManifest(
        name=f"{path.stem}-editor-preview",
        source_type=SourceType.NETWORK,
        file_name=path.name,
        required_columns=["packet_count"],
        feature_families=[FeatureFamily.NETWORK_FLOW],
        lineage={"source": "editor-preview"},
    )
    try:
        NormalizationEngine().validate_and_normalize(str(path), manifest, str(output_path))
    except (ContractValidationError, UnsupportedDatasetFormatError) as exc:
        raise DataProcessingError(str(exc), status_code=400) from exc


def _resolve_source_path(payload: dict[str, Any]) -> Path:
    if payload["dataset_format"] == DatasetFormat.PCAP.value:
        return Path(payload["projection_path"])
    return Path(payload["file_path"])


def _build_sheet_state(payload: dict[str, Any], sheet_name: str | None) -> dict[str, Any]:
    source_path = _resolve_source_path(payload)
    dataset_format = payload["dataset_format"]
    if dataset_format in {DatasetFormat.CSV.value, DatasetFormat.RES.value}:
        return _build_delimited_state(source_path, ",")
    if dataset_format == DatasetFormat.TSV.value:
        return _build_delimited_state(source_path, "\t")
    if dataset_format == DatasetFormat.JSON.value:
        return _build_json_state(source_path, payload["json_mode"])
    if dataset_format == DatasetFormat.PARQUET.value:
        return _build_parquet_state(source_path)
    if dataset_format == DatasetFormat.XLSX.value:
        if sheet_name is None:
            raise DataProcessingError("Sheet name is required for Excel editing")
        return _build_xlsx_state(source_path, sheet_name)
    if dataset_format == DatasetFormat.PCAP.value:
        return _build_delimited_state(source_path, ",")
    if dataset_format == DatasetFormat.SC.value:
        return _build_system_call_state(source_path)
    raise DataProcessingError(f"Unsupported editor format: {dataset_format}")


def _ensure_sheet_state(payload: dict[str, Any], sheet_name: str | None) -> tuple[str, dict[str, Any]]:
    active_sheet = sheet_name
    if payload["dataset_format"] == DatasetFormat.XLSX.value:
        available = payload["available_sheets"]
        if not available:
            raise DataProcessingError("Excel workbook does not contain visible sheets")
        active_sheet = sheet_name or payload.get("active_sheet") or available[0]
        if active_sheet not in available:
            raise DataProcessingError(f"Unknown sheet: {active_sheet}", status_code=404)
        payload["active_sheet"] = active_sheet
    else:
        active_sheet = None

    key = _session_sheet_key(active_sheet)
    if key not in payload["sheet_states"]:
        payload["sheet_states"][key] = _build_sheet_state(payload, active_sheet)
    return key, payload["sheet_states"][key]


def _iter_source_rows(payload: dict[str, Any], sheet_name: str | None) -> Iterator[tuple[int, dict[str, Any]]]:
    source_path = _resolve_source_path(payload)
    dataset_format = payload["dataset_format"]
    if dataset_format in {DatasetFormat.CSV.value, DatasetFormat.RES.value}:
        yield from _iter_delimited_rows(source_path, ",")
        return
    if dataset_format == DatasetFormat.TSV.value:
        yield from _iter_delimited_rows(source_path, "\t")
        return
    if dataset_format == DatasetFormat.JSON.value:
        yield from _iter_json_rows(source_path, payload["json_mode"])
        return
    if dataset_format == DatasetFormat.PARQUET.value:
        yield from _iter_parquet_rows(source_path)
        return
    if dataset_format == DatasetFormat.XLSX.value:
        if sheet_name is None:
            raise DataProcessingError("Sheet name is required for Excel editing")
        yield from _iter_xlsx_rows(source_path, sheet_name)
        return
    if dataset_format == DatasetFormat.PCAP.value:
        yield from _iter_delimited_rows(source_path, ",")
        return
    if dataset_format == DatasetFormat.SC.value:
        yield from _iter_system_call_rows(source_path)
        return
    raise DataProcessingError(f"Unsupported editor format: {dataset_format}")


def _visible_columns(state: dict[str, Any]) -> list[str]:
    deleted_columns = set(state["deleted_columns"])
    return [column for column in state["base_columns"] if column not in deleted_columns]


def _iter_visible_rows(payload: dict[str, Any], sheet_name: str | None, state: dict[str, Any]) -> Iterator[tuple[int, dict[str, Any]]]:
    deleted_rows = {int(item) for item in state["deleted_rows"]}
    visible_columns = _visible_columns(state)
    patches = state["cell_patches"]
    for row_index, row in _iter_source_rows(payload, sheet_name):
        if row_index in deleted_rows:
            continue
        row_patch = patches.get(str(row_index), {})
        values = {}
        for column in visible_columns:
            values[column] = _jsonable_value(row_patch.get(column, row.get(column)))
        yield row_index, values


def _build_session_response(payload: dict[str, Any], state: dict[str, Any]) -> EditorSessionOut:
    visible_rows = max(0, int(state["total_rows"]) - len(state["deleted_rows"]))
    total_pages = ceil(visible_rows / payload["page_size"]) if visible_rows else 0
    pending_cell_count = sum(len(item) for item in state["cell_patches"].values())
    return EditorSessionOut(
        session_id=payload["session_id"],
        file_name=payload["file_name"],
        file_path=payload["file_path"],
        dataset_format=payload["dataset_format"],
        read_only=payload["read_only"],
        page_size=payload["page_size"],
        total_rows=visible_rows,
        total_pages=total_pages,
        columns=_visible_columns(state),
        available_sheets=payload["available_sheets"],
        active_sheet=payload["active_sheet"],
        deleted_row_count=len(state["deleted_rows"]),
        deleted_columns=list(state["deleted_columns"]),
        pending_cell_count=pending_cell_count,
    )


def create_editor_session(file_path: str, page_size: int = 50, sheet_name: str | None = None) -> EditorSessionOut:
    target = _ensure_raw_file_path(file_path)
    session_id = str(uuid.uuid4())
    session_dir = get_session_dir(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    try:
        dataset_format = detect_dataset_format(target)
    except UnsupportedDatasetFormatError as exc:
        raise DataProcessingError(str(exc), status_code=400) from exc

    payload: dict[str, Any] = {
        "session_id": session_id,
        "file_path": str(target),
        "file_name": target.name,
        "dataset_format": dataset_format,
        "created_at": datetime.now(UTC).isoformat(),
        "page_size": page_size,
        "read_only": dataset_format in {DatasetFormat.PCAP.value, DatasetFormat.SC.value},
        "available_sheets": [],
        "active_sheet": None,
        "sheet_states": {},
    }

    if dataset_format == DatasetFormat.XLSX.value:
        workbook = load_workbook(target, read_only=True, data_only=False)
        try:
            payload["available_sheets"] = list(workbook.sheetnames)
        finally:
            workbook.close()
        if payload["available_sheets"]:
            payload["active_sheet"] = sheet_name or payload["available_sheets"][0]
    elif dataset_format == DatasetFormat.JSON.value:
        payload["json_mode"] = _detect_json_mode(target)
    elif dataset_format == DatasetFormat.PCAP.value:
        projection_path = session_dir / f"{target.stem}-projection.csv"
        _build_pcap_projection(target, projection_path)
        payload["projection_path"] = str(projection_path)

    _, state = _ensure_sheet_state(payload, payload["active_sheet"])
    _save_session(payload)
    return _build_session_response(payload, state)


def get_editor_page(session_id: str, page: int, sheet_name: str | None = None) -> EditorPageOut:
    payload = _load_session(session_id)
    sheet_key, state = _ensure_sheet_state(payload, sheet_name)
    response = _build_session_response(payload, state)
    if response.total_pages and page > response.total_pages:
        raise DataProcessingError(f"Requested page {page} exceeds {response.total_pages}", status_code=400)
    if not response.total_pages and page != 1:
        raise DataProcessingError("Requested page is out of range", status_code=400)

    start = (page - 1) * payload["page_size"]
    end = start + payload["page_size"]
    rows: list[EditorRowOut] = []
    for visible_index, (row_index, values) in enumerate(_iter_visible_rows(payload, payload["active_sheet"], state)):
        if visible_index < start:
            continue
        if visible_index >= end:
            break
        rows.append(EditorRowOut(row_index=row_index, values=values))

    _save_session(payload)
    return EditorPageOut(**response.model_dump(mode="json"), page=page, rows=rows)


def _require_editable_session(payload: dict[str, Any]) -> None:
    if payload["read_only"]:
        raise DataProcessingError("This file format is preview-only", status_code=409)


def _validate_row_indices(state: dict[str, Any], row_indices: list[int]) -> None:
    total_rows = int(state["total_rows"])
    for row_index in row_indices:
        if row_index < 0 or row_index >= total_rows:
            raise DataProcessingError(f"Row index out of range: {row_index}")


def update_editor_cells(session_id: str, patches: list[dict[str, Any]]) -> EditorSessionOut:
    payload = _load_session(session_id)
    _require_editable_session(payload)
    _, state = _ensure_sheet_state(payload, payload["active_sheet"])
    deleted_rows = {int(item) for item in state["deleted_rows"]}
    visible_columns = set(_visible_columns(state))

    for patch in patches:
        row_index = int(patch["row_index"])
        column = str(patch["column"])
        value = patch.get("value")
        _validate_row_indices(state, [row_index])
        if row_index in deleted_rows:
            raise DataProcessingError(f"Row {row_index} is already deleted")
        if column not in visible_columns:
            raise DataProcessingError(f"Column '{column}' is not editable")
        if not _is_scalar_value(value):
            raise DataProcessingError("Cell value must be a scalar or null")
        row_patches = state["cell_patches"].setdefault(str(row_index), {})
        row_patches[column] = value

    _save_session(payload)
    return _build_session_response(payload, state)


def delete_editor_rows(session_id: str, row_indices: list[int]) -> EditorSessionOut:
    payload = _load_session(session_id)
    _require_editable_session(payload)
    _, state = _ensure_sheet_state(payload, payload["active_sheet"])
    _validate_row_indices(state, row_indices)
    deleted_rows = {int(item) for item in state["deleted_rows"]}
    deleted_rows.update(row_indices)
    state["deleted_rows"] = sorted(deleted_rows)
    for row_index in row_indices:
        state["cell_patches"].pop(str(row_index), None)
    _save_session(payload)
    return _build_session_response(payload, state)


def delete_editor_columns(session_id: str, columns: list[str]) -> EditorSessionOut:
    payload = _load_session(session_id)
    _require_editable_session(payload)
    _, state = _ensure_sheet_state(payload, payload["active_sheet"])
    existing_columns = set(state["base_columns"])
    for column in columns:
        if column not in existing_columns:
            raise DataProcessingError(f"Unknown column: {column}")

    deleted_columns = set(state["deleted_columns"])
    deleted_columns.update(columns)
    if len(deleted_columns) >= len(state["base_columns"]):
        raise DataProcessingError("At least one column must remain after deletion")

    state["deleted_columns"] = [column for column in state["base_columns"] if column in deleted_columns]
    for row_patches in state["cell_patches"].values():
        for column in columns:
            row_patches.pop(column, None)
    _save_session(payload)
    return _build_session_response(payload, state)


def _write_delimited_file(output_path: Path, columns: list[str], rows: Iterator[dict[str, Any]], delimiter: str) -> tuple[int, int]:
    row_count = 0
    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns, delimiter=delimiter)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})
            row_count += 1
    return row_count, len(columns)


def _write_json_file(output_path: Path, columns: list[str], rows: Iterator[dict[str, Any]], mode: str) -> tuple[int, int]:
    row_count = 0
    with output_path.open("w", encoding="utf-8") as file:
        if mode == "array":
            file.write("[")
            first = True
            for row in rows:
                if not first:
                    file.write(",")
                file.write("\n  ")
                file.write(json.dumps({column: row.get(column) for column in columns}, ensure_ascii=False, default=str))
                row_count += 1
                first = False
            if row_count:
                file.write("\n")
            file.write("]")
        else:
            for row in rows:
                file.write(json.dumps({column: row.get(column) for column in columns}, ensure_ascii=False, default=str))
                file.write("\n")
                row_count += 1
    return row_count, len(columns)


def _write_xlsx_file(original_path: Path, output_path: Path, payload: dict[str, Any]) -> tuple[int, int]:
    workbook = load_workbook(original_path, read_only=True, data_only=False)
    target = Workbook(write_only=True)
    if target.worksheets:
        target.remove(target.active)

    total_rows = 0
    max_columns = 0
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = target.create_sheet(title=sheet_name)
            state = payload["sheet_states"].get(sheet_name)
            if state is None:
                source = load_workbook(original_path, read_only=True, data_only=False)
                try:
                    source_sheet = source[sheet_name]
                    for row in source_sheet.iter_rows(values_only=True):
                        worksheet.append(list(row))
                finally:
                    source.close()
                continue

            visible_columns = _visible_columns(state)
            worksheet.append(visible_columns)
            sheet_rows = 0
            for _, row in _iter_visible_rows(payload, sheet_name, state):
                worksheet.append(
                    [
                        _coerce_value(row.get(column), state["column_types"].get(column, "string"))
                        for column in visible_columns
                    ]
                )
                sheet_rows += 1
            if sheet_name == payload["active_sheet"]:
                total_rows = sheet_rows
                max_columns = len(visible_columns)
        target.save(output_path)
    finally:
        workbook.close()
        target.close()
    return total_rows, max_columns


def _arrow_value(value: Any, field_type: Any) -> Any:
    if value in ("", None):
        return None
    if pa_types.is_integer(field_type):
        return int(value)
    if pa_types.is_floating(field_type):
        return float(value)
    if pa_types.is_boolean(field_type):
        return _parse_bool(value)
    if pa_types.is_timestamp(field_type):
        return _coerce_value(value, "datetime", strict=True)
    if pa_types.is_date32(field_type) or pa_types.is_date64(field_type):
        return _coerce_value(value, "date", strict=True)
    return str(value)


def _write_parquet_file(original_path: Path, output_path: Path, payload: dict[str, Any], state: dict[str, Any]) -> tuple[int, int]:
    parquet_file = pq.ParquetFile(original_path)
    original_schema = parquet_file.schema_arrow
    visible_columns = _visible_columns(state)
    visible_schema = arrow_schema([field for field in original_schema if field.name in visible_columns])
    row_buffer: list[dict[str, Any]] = []
    row_count = 0
    writer: pq.ParquetWriter | None = None

    try:
        for _, row in _iter_visible_rows(payload, None, state):
            row_buffer.append(row)
            row_count += 1
            if len(row_buffer) >= 512:
                if writer is None:
                    writer = pq.ParquetWriter(output_path, visible_schema)
                _flush_parquet_rows(writer, visible_schema, row_buffer)
                row_buffer = []
        if writer is None:
            writer = pq.ParquetWriter(output_path, visible_schema)
        if row_buffer:
            _flush_parquet_rows(writer, visible_schema, row_buffer)
    finally:
        if writer is not None:
            writer.close()

    return row_count, len(visible_columns)


def _flush_parquet_rows(writer: pq.ParquetWriter, visible_schema: Any, rows: list[dict[str, Any]]) -> None:
    arrays = []
    for field in visible_schema:
        arrays.append(array([_arrow_value(row.get(field.name), field.type) for row in rows], type=field.type))
    writer.write_table(Table.from_arrays(arrays, schema=visible_schema))


def save_editor_session(session_id: str) -> EditorSaveOut:
    payload = _load_session(session_id)
    _require_editable_session(payload)
    sheet_key, state = _ensure_sheet_state(payload, payload["active_sheet"])
    original_path = Path(payload["file_path"])
    temp_output = get_session_dir(session_id) / f"saved-{original_path.name}"
    visible_columns = _visible_columns(state)

    def iter_rows() -> Iterator[dict[str, Any]]:
        for _, row in _iter_visible_rows(payload, payload["active_sheet"], state):
            yield {
                column: _coerce_value(row.get(column), state["column_types"].get(column, "string"))
                for column in visible_columns
            }

    dataset_format = payload["dataset_format"]
    if dataset_format in {DatasetFormat.CSV.value, DatasetFormat.RES.value}:
        row_count, column_count = _write_delimited_file(temp_output, visible_columns, iter_rows(), ",")
    elif dataset_format == DatasetFormat.TSV.value:
        row_count, column_count = _write_delimited_file(temp_output, visible_columns, iter_rows(), "\t")
    elif dataset_format == DatasetFormat.JSON.value:
        row_count, column_count = _write_json_file(temp_output, visible_columns, iter_rows(), payload["json_mode"])
    elif dataset_format == DatasetFormat.XLSX.value:
        row_count, column_count = _write_xlsx_file(original_path, temp_output, payload)
    elif dataset_format == DatasetFormat.PARQUET.value:
        row_count, column_count = _write_parquet_file(original_path, temp_output, payload, state)
    else:
        raise DataProcessingError("Saving this file format is not supported", status_code=409)

    temp_output.replace(original_path)
    stat = original_path.stat()
    saved = EditorSaveOut(
        session_id=session_id,
        file_path=str(original_path),
        size_bytes=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime, tz=UTC),
        row_count=row_count,
        column_count=column_count,
    )
    _delete_session_dir(session_id)
    return saved


def delete_editor_session(session_id: str) -> None:
    _load_session(session_id)
    _delete_session_dir(session_id)
